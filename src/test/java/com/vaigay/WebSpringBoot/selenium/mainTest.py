from time import time
from sqlUtils import getLastUser, deleteLastUser, getIdLastUser, getUserById
import openpyxl
from seleniumUtils import addUserAction, editUserAction, editUserPageAction, deleteUserAction

path = "testuser.xlsx"


def testAddUser():
    passCase = 0
    failCase = 0
    new_wb = openpyxl.load_workbook(path)

    # sheet = new_wb.active
    sheet = new_wb['AddUserTestCase']
    start = time()
    for r in range(4, 14):
        user_data = []
        for c in range(5, 11):
            cell = sheet.cell(row=r, column=c).value
            if cell is not None:
                user_data.append(cell)
            else:
                user_data.append("")
        s = time()
        mess = addUserAction(*user_data)
        expect_mess = sheet.cell(row=r, column=13).value
        if mess == expect_mess:
            if (mess == "Lưu thành công!" and tuple(user_data) == getLastUser()) or (
                    mess != "Lưu thành công!" and tuple(user_data) != getLastUser()):
                sheet.cell(row=r, column=14).value = 'PASS'
                passCase += 1
            else:
                sheet.cell(row=r, column=14).value = 'FAIL'
                failCase += 1
        else:
            sheet.cell(row=r, column=14).value = 'FAIL'
            failCase += 1
        e = time()
        print(f"{sheet.cell(row=r, column=3).value}({round(e - s, 3)})s")
        if r == 12 or r == 13:
            deleteLastUser()

    new_wb.save(path)
    end = time()
    print(f"Finish Test Add User in {round(end - start, 3)}s")
    return passCase, failCase


# testAddUser()


def testEditUser():
    passCase = 0
    failCase = 0
    new_wb = openpyxl.load_workbook(path)

    # sheet = new_wb.active
    sheet = new_wb['EditUserTestCase']
    id = getIdLastUser()
    start = time()
    for r in range(4, 10):
        user_data = []
        for c in range(5, 11):
            cell = sheet.cell(row=r, column=c).value
            if cell is not None:
                user_data.append(cell)
            else:
                user_data.append("")
        s = time()
        mess = editUserAction(id, *user_data)
        expect_mess = sheet.cell(row=r, column=13).value
        if mess == expect_mess:
            if (mess == "Sửa thành công!" and tuple(user_data) == getUserById(id)) or (
                    mess != "Sửa thành công!" and tuple(user_data) != getUserById(id)):
                sheet.cell(row=r, column=14).value = 'PASS'
                passCase += 1
            else:
                sheet.cell(row=r, column=14).value = 'FAIL'
                failCase += 1
        else:
            sheet.cell(row=r, column=14).value = 'FAIL'
            failCase += 1
        e = time()
        print(f"{sheet.cell(row=r, column=3).value}({round(e - s, 3)})s")

    new_wb.save(path)
    end = time()
    return passCase, failCase
    # print(f"Finish Test Edit User in {round(end - start, 3)}s")


def testDataEditUserPage():
    passCase = 0
    failCase = 0
    new_wb = openpyxl.load_workbook(path)

    # sheet = new_wb.active
    sheet = new_wb['EditUserTestCase']
    id = getIdLastUser()
    start = time()
    if editUserPageAction(id) == getUserById(id):
        sheet.cell(row=13, column=14).value = 'PASS'
        passCase += 1
    else:
        sheet.cell(row=13, column=14).value = 'FAIL'
        failCase += 1
    new_wb.save(path)
    end = time()
    print(f"{sheet.cell(row=13, column=3).value}({round(end - start, 3)})s")
    return passCase, failCase
    # print(f"Finish Test Edit User Data in {round(end - start, 3)}s")


def testDismissDeleteUser():
    passCase = 0
    failCase = 0
    new_wb = openpyxl.load_workbook(path)

    # sheet = new_wb.active
    sheet = new_wb['EditUserTestCase']
    id = getIdLastUser()
    start = time()
    mess = deleteUserAction(id)
    if getUserById(id):
        sheet.cell(row=16, column=14).value = 'PASS'
        passCase += 1
    else:
        sheet.cell(row=16, column=14).value = 'FAIL'
        failCase += 1
    new_wb.save(path)
    end = time()
    print(f"{sheet.cell(row=16, column=3).value}({round(end - start, 3)})s")
    return passCase, failCase


def testAcceptDeleteUser():
    passCase = 0
    failCase = 0
    new_wb = openpyxl.load_workbook(path)

    # sheet = new_wb.active
    sheet = new_wb['EditUserTestCase']
    id = getIdLastUser()
    start = time()
    delaction = deleteUserAction(id, accept=True)
    if delaction:
        sheet.cell(row=17, column=14).value = 'FAIL'
        failCase += 1
    else:
        sheet.cell(row=17, column=14).value = 'PASS'
        passCase += 1
    new_wb.save(path)
    end = time()
    print(f"{sheet.cell(row=17, column=3).value}({round(end - start, 3)})s")
    return passCase, failCase


totalTest = {
    'passCase': 0,
    'failCase': 0
}
passCase, failCase = testAddUser()
totalTest['passCase'] += passCase
totalTest['failCase'] += failCase
print("==============================")
start = time()
passCase, failCase = testEditUser()
totalTest['passCase'] += passCase
totalTest['failCase'] += failCase
passCase, failCase = testDataEditUserPage()
totalTest['passCase'] += passCase
totalTest['failCase'] += failCase
end = time()
print(f"Finish Test Edit User Data in {round(end - start, 3)}s")
print("==============================")
start = time()
passCase, failCase = testDismissDeleteUser()
totalTest['passCase'] += passCase
totalTest['failCase'] += failCase
passCase, failCase = testAcceptDeleteUser()
totalTest['passCase'] += passCase
totalTest['failCase'] += failCase
end = time()
print(f"Finish Test Delete User Data in {round(end - start, 3)}s")
print("==============================")
print(f"TestCase: {totalTest['passCase'] + totalTest['failCase']}")
print(f"Pass: {totalTest['passCase']}")
print(f"Fail: {totalTest['failCase']}")
